"""
Utilitários gerais do sistema
"""

import qrcode
import io
import base64
from PIL import Image
from django.http import HttpResponse
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class QRCodeGenerator:
    """
    Gerador de QR Codes para o sistema
    """
    
    @staticmethod
    def generate_line_qr_code(production_line, format='PNG'):
        """
        Gera QR Code para linha de produção
        """
        # Dados do QR Code
        qr_data = {
            'type': 'production_line',
            'line_code': production_line.code,
            'line_name': production_line.name,
            'url': f"/line-summary/{production_line.code}/",
            'timestamp': timezone.now().isoformat()
        }
        
        # Criar string do QR Code
        qr_string = f"Linha: {qr_data['line_name']}\n"
        qr_string += f"Código: {qr_data['line_code']}\n"
        qr_string += f"Acesso: {qr_data['url']}\n"
        qr_string += f"Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        return QRCodeGenerator._create_qr_image(qr_string, format)
    
    @staticmethod
    def generate_shift_qr_code(production_line, shift, date, format='PNG'):
        """
        Gera QR Code para turno específico
        """
        qr_data = {
            'type': 'shift_summary',
            'line_code': production_line.code,
            'shift': shift.name,
            'date': date.strftime('%Y-%m-%d'),
            'url': f"/shift-summary/{date.strftime('%Y-%m-%d')}/{shift.name}/{production_line.id}/",
        }
        
        qr_string = f"Turno: {shift.get_name_display()}\n"
        qr_string += f"Linha: {production_line.name}\n"
        qr_string += f"Data: {date.strftime('%d/%m/%Y')}\n"
        qr_string += f"Acesso: {qr_data['url']}"
        
        return QRCodeGenerator._create_qr_image(qr_string, format)
    
    @staticmethod
    def generate_report_qr_code(quality_report, format='PNG'):
        """
        Gera QR Code para laudo de qualidade
        """
        qr_string = f"Laudo: {quality_report.report_number}\n"
        qr_string += f"Produto: {quality_report.product.name}\n"
        qr_string += f"Status: {quality_report.get_status_display()}\n"
        qr_string += f"Acesso: /report/{quality_report.id}/"
        
        return QRCodeGenerator._create_qr_image(qr_string, format)
    
    @staticmethod
    def _create_qr_image(data, format='PNG'):
        """
        Cria imagem do QR Code
        """
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        
        # Criar imagem
        img = qr.make_image(fill_color="black", back_color="white")
        
        if format.upper() == 'BASE64':
            # Retornar como base64 para uso em templates
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            return f"data:image/png;base64,{img_base64}"
        else:
            # Retornar imagem PIL
            return img


class DataExporter:
    """
    Exportador de dados para Excel e CSV
    """
    
    @staticmethod
    def export_spot_analyses_to_excel(queryset, filename=None):
        """
        Exporta análises pontuais para Excel
        """
        if not filename:
            filename = f"analises_pontuais_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Análises Pontuais"
        
        # Cabeçalhos
        headers = [
            'Data', 'Turno', 'Linha', 'Produto', 'Propriedade', 
            'Sequência', 'Valor', 'Unidade', 'Status', 'Hora', 
            'Operador', 'Ação Tomada'
        ]
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for row, analysis in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=analysis.date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=analysis.shift.get_name_display())
            ws.cell(row=row, column=3, value=analysis.production_line.name)
            ws.cell(row=row, column=4, value=analysis.product.name)
            ws.cell(row=row, column=5, value=f"{analysis.property.identifier} - {analysis.property.name}")
            ws.cell(row=row, column=6, value=analysis.sequence)
            ws.cell(row=row, column=7, value=float(analysis.value))
            ws.cell(row=row, column=8, value=analysis.property.unit or '')
            ws.cell(row=row, column=9, value=analysis.get_status_display())
            ws.cell(row=row, column=10, value=analysis.sample_time.strftime('%H:%M'))
            ws.cell(row=row, column=11, value=analysis.operator.get_full_name() or analysis.operator.username)
            ws.cell(row=row, column=12, value=analysis.action_taken or '')
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer, filename
    
    @staticmethod
    def export_quality_reports_to_excel(queryset, filename=None):
        """
        Exporta laudos de qualidade para Excel
        """
        if not filename:
            filename = f"laudos_qualidade_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laudos de Qualidade"
        
        headers = [
            'Número do Laudo', 'Tipo', 'Produto', 'Linha', 'Data Inicial', 
            'Data Final', 'Cliente', 'Lote', 'Quantidade (ton)', 'Status',
            'Criado em', 'Criado por', 'Aprovado por', 'Aprovado em'
        ]
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row, report in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=report.report_number)
            ws.cell(row=row, column=2, value=report.get_report_type_display())
            ws.cell(row=row, column=3, value=report.product.name)
            ws.cell(row=row, column=4, value=report.production_line.name)
            ws.cell(row=row, column=5, value=report.start_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=6, value=report.end_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=7, value=report.customer_name or '')
            ws.cell(row=row, column=8, value=report.batch_number or '')
            ws.cell(row=row, column=9, value=float(report.quantity) if report.quantity else '')
            ws.cell(row=row, column=10, value=report.get_status_display())
            ws.cell(row=row, column=11, value=report.created_at.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=12, value=report.created_by.get_full_name() or report.created_by.username)
            ws.cell(row=row, column=13, value=report.approved_by.get_full_name() if report.approved_by else '')
            ws.cell(row=row, column=14, value=report.approved_at.strftime('%d/%m/%Y %H:%M') if report.approved_at else '')
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer, filename
    
    @staticmethod
    def export_to_csv(queryset, fields, filename=None):
        """
        Exporta queryset para CSV
        """
        if not filename:
            filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Adicionar BOM para Excel reconhecer UTF-8
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Cabeçalhos
        headers = []
        for field in fields:
            if hasattr(queryset.model, field):
                field_obj = queryset.model._meta.get_field(field)
                headers.append(field_obj.verbose_name)
            else:
                headers.append(field)
        
        writer.writerow(headers)
        
        # Dados
        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field, '')
                if hasattr(value, 'strftime'):  # Data/datetime
                    value = value.strftime('%d/%m/%Y %H:%M') if hasattr(value, 'hour') else value.strftime('%d/%m/%Y')
                elif hasattr(value, '__call__'):  # Métodos
                    value = value()
                row.append(str(value))
            writer.writerow(row)
        
        return response


class DataImporter:
    """
    Importador de dados históricos
    """
    
    @staticmethod
    def import_from_excel(file_path, sheet_name=None):
        """
        Importa dados de arquivo Excel
        """
        try:
            # Ler arquivo Excel
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            return df, None
        except Exception as e:
            return None, str(e)
    
    @staticmethod
    def import_spot_analyses_from_dataframe(df, user, validate=True):
        """
        Importa análises pontuais de DataFrame
        """
        from quality_control.models import SpotAnalysis, Product, Property
        from core.models import ProductionLine, Shift
        
        imported_count = 0
        errors = []
        
        required_columns = ['data', 'turno', 'linha', 'produto', 'propriedade', 'valor']
        
        # Verificar colunas obrigatórias
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return 0, [f"Colunas obrigatórias ausentes: {', '.join(missing_columns)}"]
        
        for index, row in df.iterrows():
            try:
                # Validar e buscar objetos relacionados
                date = pd.to_datetime(row['data']).date()
                
                # Buscar turno
                shift = Shift.objects.filter(name=row['turno']).first()
                if not shift:
                    errors.append(f"Linha {index + 2}: Turno '{row['turno']}' não encontrado")
                    continue
                
                # Buscar linha de produção
                line = ProductionLine.objects.filter(
                    models.Q(name=row['linha']) | models.Q(code=row['linha'])
                ).first()
                if not line:
                    errors.append(f"Linha {index + 2}: Linha de produção '{row['linha']}' não encontrada")
                    continue
                
                # Buscar produto
                product = Product.objects.filter(
                    models.Q(name=row['produto']) | models.Q(code=row['produto'])
                ).first()
                if not product:
                    errors.append(f"Linha {index + 2}: Produto '{row['produto']}' não encontrado")
                    continue
                
                # Buscar propriedade
                property_obj = Property.objects.filter(
                    models.Q(name=row['propriedade']) | models.Q(identifier=row['propriedade'])
                ).first()
                if not property_obj:
                    errors.append(f"Linha {index + 2}: Propriedade '{row['propriedade']}' não encontrada")
                    continue
                
                # Sequência (padrão 1 se não informado)
                sequence = row.get('sequencia', 1)
                
                # Criar análise pontual
                analysis = SpotAnalysis(
                    date=date,
                    shift=shift,
                    production_line=line,
                    product=product,
                    property=property_obj,
                    sequence=sequence,
                    value=row['valor'],
                    operator=user,
                    action_taken=row.get('acao_tomada', '')
                )
                
                if validate:
                    analysis.full_clean()
                
                analysis.save()
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Linha {index + 2}: {str(e)}")
        
        return imported_count, errors


class NotificationService:
    """
    Serviço de notificações do sistema
    """
    
    @staticmethod
    def send_approval_notification(quality_report, approved_by):
        """
        Envia notificação de aprovação de laudo
        """
        # Implementar integração com email/SMS/Slack
        # Por enquanto, apenas log
        print(f"Laudo {quality_report.report_number} aprovado por {approved_by.username}")
    
    @staticmethod
    def send_alert_notification(spot_analysis):
        """
        Envia notificação de análise em alerta
        """
        if spot_analysis.status in ['ALERT', 'REJECTED']:
            print(f"ALERTA: Análise {spot_analysis.property.identifier} fora dos limites - Valor: {spot_analysis.value}")
    
    @staticmethod
    def send_daily_summary(date=None):
        """
        Envia resumo diário
        """
        if not date:
            date = timezone.now().date()
        
        from quality_control.models import SpotAnalysis
        
        total_analyses = SpotAnalysis.objects.filter(date=date).count()
        alerts = SpotAnalysis.objects.filter(date=date, status__in=['ALERT', 'REJECTED']).count()
        
        print(f"Resumo do dia {date.strftime('%d/%m/%Y')}: {total_analyses} análises, {alerts} alertas")


class AuditLogger:
    """
    Logger de auditoria do sistema
    """
    
    @staticmethod
    def log_user_action(user, action, object_type, object_id, details=None):
        """
        Registra ação do usuário
        """
        from core.models import AuditLog
        
        AuditLog.objects.create(
            user=user,
            action=action,
            object_type=object_type,
            object_id=object_id,
            details=details or {},
            ip_address=None,  # Implementar captura de IP se necessário
            user_agent=None   # Implementar captura de User-Agent se necessário
        )
    
    @staticmethod
    def log_report_approval(quality_report, approved_by):
        """
        Registra aprovação de laudo
        """
        AuditLogger.log_user_action(
            user=approved_by,
            action='APPROVE_REPORT',
            object_type='QualityReport',
            object_id=quality_report.id,
            details={
                'report_number': quality_report.report_number,
                'product': quality_report.product.name,
                'approved_at': timezone.now().isoformat()
            }
        )
    
    @staticmethod
    def log_data_export(user, export_type, record_count):
        """
        Registra exportação de dados
        """
        AuditLogger.log_user_action(
            user=user,
            action='EXPORT_DATA',
            object_type=export_type,
            object_id=None,
            details={
                'record_count': record_count,
                'exported_at': timezone.now().isoformat()
            }
        )
